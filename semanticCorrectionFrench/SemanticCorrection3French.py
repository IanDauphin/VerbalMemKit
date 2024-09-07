from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict


# Read the Excel file (replace if using different file path)
excel_file_path = "/Users/ianda_hpfqchp/Library/CloudStorage/OneDrive-UniversitédeMoncton/Dominic OSF Data/Special Characters Debug.xlsx"
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Define the fill color for highlighted cells (brown)
fill_brown = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
fill_grey = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

# Function to check if a string contains any vowels, later used to check vowelless words
def contains_vowels(text):
    vowels = set('aeiouy')
    return any(char in vowels for char in text)

# Function to check for special characters in a string
def has_special_characters(string):
    special_characters = "[].,'\\/?:;-*~"  # Add more if needed
    return any(char in string for char in special_characters)

# Function to check for three adjacent instances of the same character in a string
def has_three_adjacent_characters(string):
    for i in range(len(string) - 2):
        if string[i] == string[i+1] == string[i+2]:
            return True
    return False

def generate_inserted_permutations(word):
    permutations_list = set()
    for i in range(len(word) + 1):
        for letter in "aâàäbcdeéèêëfghiîïjklmnôöopqrstuûüvwxyz":
            new_permutation = word[:i] + letter + word[i:]
            permutations_list.add(new_permutation)
    return permutations_list

def generate_deleted_permutations(word):
    permutations_list = set()
    for i in range(len(word)):
        new_permutation = word[:i] + word[i+1:]
        permutations_list.add(new_permutation)
    return permutations_list

def generate_substituted_permutations(word):
    alphabet = "abcdefghijklmnopqrstuvwxyz"
    permutations_list = set()
    for i in range(len(word)):
        for letter in alphabet:
            if letter != word[i]:
                new_permutation = word[:i] + letter + word[i+1:]
                permutations_list.add(new_permutation)
    return permutations_list

def generate_special_permutations(word, special_characters):
    permutations_list = set()

    # Generate permutations by adding special characters
    for i in range(len(word) + 1):
        for char in special_characters:
            new_permutation = word[:i] + char + word[i:]
            permutations_list.add(new_permutation)

    # Generate permutations by swapping with special characters
    for i in range(len(word)):
        for char in special_characters:
            # Check if the character requires a tilde in Excel
            new_permutation = word[:i] + char + word[i+1:]
            permutations_list.add(new_permutation)

    return permutations_list

def generate_all_permutations(word, special_characters= "[].,'\\/?:;-*"):
    inserted_permutations = generate_inserted_permutations(word)
    deleted_permutations = generate_deleted_permutations(word)
    substituted_permutations = generate_substituted_permutations(word)
    special_permutations = generate_special_permutations(word, special_characters)
    word_itself = {word}

    all_permutations = word_itself | inserted_permutations | deleted_permutations |  substituted_permutations | special_permutations 
    return all_permutations

def find_shared_permutations(word_i, word_j):
    permutations_i = generate_all_permutations(word_i)
    permutations_j = generate_all_permutations(word_j)

    common_permutations = permutations_i & permutations_j
    return common_permutations

def highlight_cells(sheet, row_index, column_indices, fill_color):
    for col_index in column_indices:
        sheet.cell(row=row_index, column=col_index).fill = fill_color


# Define columns for responses and correct words
responses_columns = "K:P" # Replace with appropriate columns
correct_words_columns = "E:J" # Replace with appropriate columns

# Get column indices for responses and correct words
responses_column_indices = [col[0].column for col in sheet[responses_columns]]
correct_words_column_indices = [col[0].column for col in sheet[correct_words_columns]]


# Iterate through each row to apply conditional formatting
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    for col_index in range(responses_column_indices[0], responses_column_indices[-1] + 1):
        response = str(row[col_index - 1]).lower()
        correct_word = str(row[correct_words_column_indices[col_index - responses_column_indices[0]] - 1])


        # Check if the cell value meets the specified conditions
        if response in ["something", "somethin", "soemthings", "somethings", "gg", "nn", "pass", "skip'", "'skip", "skip", "sskip", "skkip", "skiip", "skipp", "lol", "lmao", "\\", "/", "don't", "'dont", "don;t", "don:t", "dont", "idontk", "know", "don't know", "don'tknow", "forgot", "forgotten", "dontknow", "dunno",  "don’t know", "-", "unsure", "forgot", "x", " ", "   ",  "?", "'?", "n/a", "'n\\a", "n/an/a", "nn/a", "n/aa" "n/", "/a", "na", "na", "n\a", "n\\a", "n/s", "n/z", "ns", "n/q", "nd", "nq", "unknown", "'unknown", "nknown", "unnknown", "uuknown", "unknow", "ounknown", "unknovwn", "unknfown", "unknownm", "ubknown", "unknovn", "unkvnown", "uknown", "unknowsn", "unknowgn", "unkvown", "unknonwn", "unknodwn", "unknqwn", "unknowz", "unknowa", "unsknown", "unknowjn", "unpnown", "unkbown", "unknohn", "unknowqn", "unknopn", "unkenown", "unknomwn", "unknhwn", "unknxown", "unknowny", "dunknown", "unknonn", "inknown", "unkonown", "unknozwn", "unknowrn", "unkcown", "unknowt", "uxnknown", "unknoyn", "unvknown", "unktown", "unknownl", "unfnown", "nunknown", "unknodn", "unknocwn", "unkoown", "unkfown", "uvnknown", "punknown", "unknoan", "unknowyn", "unknnwn", "unkwnown", "uqknown", "ungnown", "unknqown", "unkyown", "unpknown", "unknawn", "unknyown", "unknvwn", "uynknown", "unkniown", "unknotn", "unknocn", "unknowan", "unknjown", "unknorn", "unkiown", "untknown", "unknowmn", "unknowo", "unknowhn", "unknownc", "upknown", "ulknown", "unknoww", "uanknown", "unklnown", "unknowun", "unknolwn", "uneknown", "unknowen", "unknownh", "unyknown", "unnnown", "unrknown", "unnknown", "unkxown", "ynknown", "unknotwn", "unknowvn", "unkneown", "undnown", "unknopwn", "unknsown", "unknoun", "onknown", "unksnown", "unknownp", "unktnown", "unwknown", "unjnown", "unkndown", "unkjown", "unknwwn", "unknownv", "unkpown", "uncnown", "cnknown", "unknow", "unoknown", "unknobn", "unknpwn", "unkqnown", "lnknown", "unhnown", "uenknown", "unknowy", "unkkown", "unknojwn", "ueknown", "unknon", "uuknown", "unkngown", "unknownf", "unkntwn", "ufnknown", "undknown", "unkaown", "unknowm", "unkmown", "unknozn", "usknown", "unenown", "dnknown", "unfknown", "unknwn", "unknoen", "unknoon", "unknywn", "unkncwn", "unknowj", "unkrnown", "munknown", "kunknown", "uaknown", "unknogn", "unknownt", "unmknown", "tunknown", "unknowq", "unkjnown", "unqknown", "unknofwn", "unknuwn", "ungknown", "uwknown", "unkuown", "ujnknown", "uninown", "ujknown", "unknowe", "unknnown", "unknlown", "unkcnown", "gnknown", "snknown", "pnknown", "upnknown", "cunknown", "funknown", "nnknown", "uknknown", "uhnknown", "unwnown", "umnknown", "unknmown", "qunknown", "uoknown", "unknoswn", "unknorwn", "unknowwn", "hunknown", "unknuown", "unknowd", "aunknown", "uonknown", "xnknown", "unkgown", "wnknown", "bnknown", "vnknown", "ugnknown", "ulnknown", "unkgnown", "unknokwn", "unkown", "unknowv", "unvnown", "unknowk", "mnknown", "unknoxwn", "unbknown", "udknown", "unklown", "unknvown", "unknlwn", "ugknown", "unknowtn", "unknomn", "unknowg", "unknzwn", "unknowdn", "unkbnown", "unknownj", "nknown", "unknownz", "unqnown", "unknowr", "unkwown", "ufknown", "unknofn", "unzknown", "udnknown", "unknowon", "iunknown", "bunknown", "unsnown", "unknoewn", "sunknown", "unknwown", "enknown", "uvknown", "unhknown", "unknowin", "unkeown", "unknows", "unkinown", "unknoown", "unknowpn", "unknownr", "unaknown", "unknowx", "urnknown", "unnown", "ucnknown", "unknobwn", "uhknown", "unxnown", "unkdown", "uncknown", "unknownx", "ucknown", "unknswn", "runknown", "unknowc", "unknownb", "unknmwn", "unkdnown", "unknowu", "uqnknown", "unjknown", "unlknown", "unknohwn", "uzknown", "unkknown", "unkhnown", "hnknown", "unkunown", "unknowzn", "unynown", "unknowl", "ubnknown", "uxknown", "unknowbn", "unknowh", "unknowcn", "unknownk", "unknpown", "unkqown", "unrnown", "unonown", "tnknown", "untnown", "unznown", "unknjwn", "unkncown", "unknowln", "unknhown", "unknewn", "unmnown", "fnknown", "unknownw", "uznknown", "uyknown", "unkpnown", "unknaown", "unknowkn", "unksown", "unanown", "unknownu", "junknown", "knknown", "umknown", "unknoxn", "urknown", "unknouwn", "unknojn", "unknowxn", "unknoqn", "lunknown", "unkznown", "unkhown", "unknzown", "unknowb", "unknrown", "unknowns", "unkrown", "unknbown", "xunknown", "unkanown", "unknownd", "unkngwn", "uniknown", "znknown", "unbnown", "unuknown", "uiknown", "jnknown", "ununown", "unkndwn", "unknownn", "unknowna", "utnknown", "unknogwn", "unknowni", "unkniwn", "unkynown", "unknosn", "unknbwn", "unknkwn", "unkfnown", "unknowno", "unknowfn", "ukknown", "yunknown", "unkzown", "unknoawn", "unkxnown", "unknoywn", "unlnown", "unknkown", "uunknown", "unknfwn", "unknowne", "unknxwn", "unknoqwn", "uinknown", "unknoin", "anknown", "uwnknown", "unxknown", "vunknown", "unknowp", "wunknown", "zunknown", "rnknown", "qnknown", "unkmnown", "unknownq", "unknowi", "usnknown", "unknowf", "eunknown", "unknoln", "unkntown", "unknowng", "utknown", "unknrwn", "gunknown", "unknokn", "unknoiwn",  ".", "ok", "??", "?" "dk", "idk", "no", "'notsure", "notsure'", "'unsure", "unsure'", "not sure", "notsure", "notsur", "otsure", "nnotsure", "nootsure", "nottsure", "notssure", "sotsuure", "notsurre", "notsuree", "noidea", "notaclue", "noclue", "a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z", "idontknow", "cant", "remember", "words", "unsuer", "unsue", "ahhh", "uhhh"] :
            sheet.cell(row=row_index, column=col_index).value = 0
            sheet.cell(row=row_index, column=col_index).fill = fill_red

# Iterate through each row to check for invalid responses
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        responses = [str(row[col - 1]) for col in responses_column_indices]
        # Define the header for the new column in AN (replace column if header different length)
        new_column_header = "Trial Validity"
        # Get the last column index in the responses columns
        last_responses_column_index = responses_column_indices[-1]
        # Calculate the column index for the new column
        new_column_index = last_responses_column_index + 2
        # Insert the new column header in the first row
        sheet.cell(row=1, column=new_column_index, value=new_column_header)

        # Check if three cells in responses have the same contents and that content is not "0"
        invalid_count = 0
        for response in set(responses):
            if responses.count(response) > 2 and response not in ["0", "na"]:
                invalid_count += 1

        # Update the cell in column "AN" based on the condition 
        if invalid_count > 0:
            sheet.cell(row=row_index, column=18, value="INVALID")
            sheet.cell(row=row_index, column=18,).fill = fill_red

        else:
            sheet.cell(row=row_index, column=18, value="valid")
            sheet.cell(row=row_index, column=18,).fill = fill_green

# Dictionaries to store unique and shared permutations for each current_correct_word
unique_permutations_dict = {}
shared_permutations_dict = {}
spam_responses_dict = {} # this line most likely doesnt do anything but I ain't touching it for now

# Iterate through each row in the Excel file
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    responses = [str(row[col - 1]) for col in responses_column_indices]
    correct_words = [str(row[col - 1]) for col in correct_words_column_indices]

    # Dictionary to store multiple responses for the same row and same current_correct_word
    current_row_responses_dict = {}


    # Iterate over each correct word
    for i in range(len(correct_words)):
        current_correct_word = correct_words[i]

        # Find all possible permutations for the current correct word
        all_permutations_current = generate_all_permutations(current_correct_word)


        # Compare with other correct words
        for j in range(len(correct_words)):
            if i != j:  # Exclude comparing with itself
                other_correct_word = correct_words[j]

                # Find permutations unique to the current correct word
                unique_permutations = all_permutations_current - find_shared_permutations(current_correct_word, other_correct_word)

                # Store unique permutations in the dictionary
                unique_permutations_dict.setdefault(current_correct_word, []).extend(unique_permutations)

                # Find shared permutations
                shared_permutations = find_shared_permutations(current_correct_word, other_correct_word)

                # Store shared permutations in the dictionary
                shared_permutations_dict.setdefault(current_correct_word, []).extend(shared_permutations)

                # Highlight cells with shared permutations in brown
                for col_index, response in enumerate(responses):
                    if response in shared_permutations:
                        sheet.cell(row=row_index, column=responses_column_indices[col_index]).fill = fill_brown

    # Object to store responses and their corresponding correct words
    responses_object = {}

    # Store all multiple responses for the same row and same current_correct_word before corrections
    for col_index, response in enumerate(responses):
        for i in range(len(correct_words)):
            current_correct_word = correct_words[i]
            current_row_responses_dict.setdefault((row_index, current_correct_word), []).append(response)

    # Check if any value in the responses columns is in the dictionary of unique permutations
    # for one of the correct words and not in any shared dictionary for that correct word in the same row
    for col_index, response in enumerate(responses):
        for i in range(len(correct_words)):
            current_correct_word = correct_words[i]
            

            if (
                response in unique_permutations_dict.get(current_correct_word, []) and
                response not in shared_permutations_dict.get(current_correct_word, [])
                and (row_index, current_correct_word) not in spam_responses_dict # this line most likely doesnt do anything but I ain't touching it for now
            ):
                
                # Add the response to the responses_object
                responses_object.setdefault((row_index, current_correct_word), []).append(response)
    
    for col_index, response in enumerate(responses):
        for i in range(len(correct_words)):
            current_correct_word = correct_words[i]


        # Check the responses_object and correct responses based on the condition
        for (row_index, current_correct_word), response_list in responses_object.items():
            unique_values_responses1 = set() # Initialize the set for each row
            corrected = False  # Flag to track if the first occurrence has been corrected, if it is a permutation of the corresponding correct word

            # Correct the responses to their corresponding correct word for the first occurrence
            for col_index, response in enumerate(responses):
                if response in response_list and not corrected:
                    sheet.cell(row=row_index, column=responses_column_indices[col_index]).value = current_correct_word #if first occurance, correct with corresponding correct word
                    # Add the response to the set for the current row
                    unique_values_responses1.add(response)

                    corrected = True  # Set the flag to True after correcting the first occurrence

                elif response in response_list and response in unique_values_responses1:
                    #Replace subsequent occurrences with 0
                    sheet.cell(row=row_index, column=responses_column_indices[col_index]).fill = fill_yellow #yellow cells are other occurance of word after first (whether different typos or not)

# Iterate through each row to apply conditional formatting
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    for col_index in range(responses_column_indices[0], responses_column_indices[-1] + 1):
        response = str(row[col_index - 1]).lower()
        correct_word = str(row[correct_words_column_indices[col_index - responses_column_indices[0]] - 1])

        # Check if the cell value meets the specified conditions
        if response == "0" :
            sheet.cell(row=row_index, column=col_index).fill = fill_red

# Dictionary to store participant numbers and their responses
participant_responses = defaultdict(list)

# Iterate through each row to apply conditional formatting
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    subject_number = row[0]  # Assuming subject number is in the first column
    responses = [str(cell) for cell in row[responses_column_indices[0]-1:responses_column_indices[-1]]]

    # Define the header for the new column in AN (replace column if header different length)
    new_column_header2 = "Compliance"
    # Get the last column index in the responses columns
    last_responses_column_index = responses_column_indices[-1]
    # Calculate the column index for the new column
    new_column_index = last_responses_column_index + 3
    # Insert the new column header in the first row
    sheet.cell(row=1, column=new_column_index, value=new_column_header2)

    # Update participant_responses dictionary
    participant_responses[subject_number].extend(responses)

# Iterate through participant responses to track words appearing more than three times
for subject_number, responses in participant_responses.items():
    word_count = defaultdict(int)
    for response in responses:
        if response != '' and response != "0":
            word_count[response.lower()] += 1

    # Find the row indices corresponding to the subject number
    subject_row_indices = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == subject_number:
                subject_row_indices.append(row_index)

    # Set the value for all rows corresponding to the subject number based on word count
    for subject_row_index in subject_row_indices:
        if any(count > 3 for count in word_count.values()):
            sheet.cell(row=subject_row_index, column=19, value="NO").fill = fill_red
        else:
            sheet.cell(row=subject_row_index, column=19, value="yes").fill = fill_green

    # Highlight words appearing more than three times within the response columns
    for word, count in word_count.items():
        for subject_row_index in subject_row_indices:
            for col_index in responses_column_indices:  # Iterate only over response columns
                response = str(sheet.cell(row=subject_row_index, column=col_index).value).lower()
                if count >= 3 and word == response:
                    sheet.cell(row=subject_row_index, column=col_index).fill = fill_orange

# Iterate through each row to apply conditional formatting
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    unique_values_responses2 = set()  # Initialize the set for each row

    for col_index in range(responses_column_indices[0], responses_column_indices[-1] + 1):
        response = str(row[col_index - 1]).lower()
        correct_word = str(row[correct_words_column_indices[col_index - responses_column_indices[0]] - 1])

        # Here I am checking if the response contains no vowels. If not, its probably garbage.
        if response and response != "0" and (not contains_vowels(response) or has_special_characters(response) or has_three_adjacent_characters(response)):
            sheet.cell(row=row_index, column=col_index).fill = fill_grey  # Highlight cell with special characters

        # Add your condition for formatting (replace it with your actual condition)
        if response == correct_word:
            sheet.cell(row=row_index, column=col_index).fill = fill_green
        
        if response in unique_values_responses2 and response != "0":
            sheet.cell(row=row_index, column=col_index).fill = fill_yellow
            
            
        # Add the response to the set for the current row
        unique_values_responses2.add(response)

        


# Save the modified Excel file
workbook.save("/Users/ianda_hpfqchp/Library/CloudStorage/OneDrive-UniversitédeMoncton/Dominic OSF Data/Corrected - Special Characters Debug.xlsx")
